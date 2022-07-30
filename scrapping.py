from bs4 import BeautifulSoup
import requests
import csv
from openpyxl import load_workbook

headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'}

# ilk 1566 markanın ismi
url = 'https://www.sikayetvar.com/tum-markalar'
source = requests.get(url, verify=False, headers=headers)
soup = BeautifulSoup(source.content)
companies = soup.find('ul',class_='brand-list').find_all('li')

# ilk 1566 markanın ismi -1.sayfadakiler
company_ratio = {}
for c in companies:
    # company_ratio[c.find('span',class_='brand-name').text] = int(c.find('strong').text.replace('.',''))
    company_ratio['https://www.sikayetvar.com'+c.a['href']] = int(c.find('strong').text.replace('.',''))

# ilk 1566 markanın ismi -2.sayfa ve sonrakindekiler
for page_idx in range(2,100):
    url = 'https://www.sikayetvar.com/tum-markalar?page='+str(page_idx)
    source = requests.get(url, verify=False, headers=headers)
    soup = BeautifulSoup(source.content)
    companies = soup.find('ul',class_='brand-list').find_all('li')
    for c in companies:
        company_ratio['https://www.sikayetvar.com'+c.a['href']] = int(c.find('strong').text.replace('.',''))

# ilk 1566 markanın ana sayfalarındaki toplam 23127 sikayet linkleri
link_list = []
for url in company_ratio.keys():
    source = requests.get(url, verify=False, headers=headers)
    soup = BeautifulSoup(source.content)
    links = soup.find_all('h2',class_='complaint-title')
    for l in links:
        link = 'https://www.sikayetvar.com'+l.a.get('href')
        link_list.append(link)

# verinin saklanacağı dosyaya baslıklar yazılıyor
filename = 'sikayetvar_data.xlsx'
wb = load_workbook(filename = filename)
sheets = wb.sheetnames
ws = wb[sheets[0]]
values = ['URL','COMPANY','TITLE','POST_TIME','COMPLAINT']
excel_row = 1
for col in range(1,6):
    ws.cell(column =col,row=excel_row).value = values[col-1]
excel_row = excel_row+1

# sikayetler
for url in link_list:
    # degerlerin degiskenlere aktarılması
    source = requests.get(url, verify=False, headers=headers)
    soup = BeautifulSoup(source.content)
    
    try:
        company = soup.find('ul',class_='breadcrumb').find_all('a')[1]['title']
    except Exception as e:
        company = ''
    # print(company)
    
    try:
        title = soup.find('div',class_='story-title').h1.text
    except Exception as e:
        title = ''
    # print(title)
    
    try:
        post_time = soup.find('span',class_='post-time')['title']
    except Exception as e:
        post_time = ''
    # print(post_time)
    
    try:
        complaint_container = soup.find('div',class_='card-text').find_all('p')
        complaint = ''
        for c in complaint_container:
            complaint = complaint+''+c.text+'\n'
    except Exception as e:
        complaint = ''
    # print(complaint.strip())
    
    
    # verinin saklanacağı dosyaya degerler yazılıyor
    values = [url,company,title,post_time,complaint.strip()]
    for col in range(1,6):
        ws.cell(column =col,row=excel_row).value = values[col-1]
    excel_row = excel_row+1

wb.save(filename)

